# productos/views_admin.py
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, F, Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta

from .models import Producto, Categoria, Promocion, Promotion
from .forms import ProductoForm, CategoriaForm
from usuarios.decorators import group_required
# productos/views_admin.py
from .forms import ProductoForm, CategoriaForm
from usuarios.models import Usuario, Rol
from usuarios.forms import UsuarioForm
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.contrib.auth.models import Group



# --------- Dashboard ---------
@login_required
@group_required("Admin", "Empleado")
def dashboard(request):
    from django.utils import timezone
    from django.db.models import Sum, Count, Q
    from datetime import timedelta
    from ventas.models import Venta, VentaDetalle
    
    # Obtener fechas del request (filtros personalizados)
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    
    hoy = timezone.localdate()
    
    # Parsear fechas si se proporcionan
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = timezone.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = timezone.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_inicio = hoy - timedelta(days=30)
            fecha_fin = hoy
    else:
        # Default: últimos 30 días
        fecha_inicio = hoy - timedelta(days=30)
        fecha_fin = hoy
    
    # Calcular hace_7_dias desde fecha_fin
    hace_7_dias = fecha_fin - timedelta(days=7)
    
    # Convertir dates a datetimes para filtrado correcto con timezones
    dt_inicio = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
    dt_fin = timezone.make_aware(datetime.combine(fecha_fin, datetime.max.time()))
    dt_hace_7 = timezone.make_aware(datetime.combine(hace_7_dias, datetime.min.time()))
    
    # ===== INVENTARIO =====
    low_stock = Producto.objects.filter(stock_actual__lte=F("stock_minimo")).select_related("categoria")
    total_prod = Producto.objects.count()
    stock_total = Producto.objects.aggregate(total=Sum('stock_actual'))['total'] or 0
    productos_sin_stock = Producto.objects.filter(stock_actual=0).count()
    
    # Productos por categoría (para gráfico)
    productos_por_categoria = list(
        Categoria.objects.annotate(cantidad=Count('producto')).values('nombre', 'cantidad').order_by('-cantidad')
    )
    
    # ===== RENTABILIDAD POR CATEGORÍA =====
    rentabilidad_categoria = list(
        VentaDetalle.objects.filter(venta__creado_en__gte=dt_inicio, venta__creado_en__lte=dt_fin)
        .values('producto__categoria__nombre')
        .annotate(
            total_vendido=Sum('cantidad'),
            ingresos=Sum('subtotal'),
            num_ventas=Count('venta', distinct=True)
        )
        .order_by('-ingresos')
    )
    
    # ===== VENTAS =====
    # Ventas en el rango de fechas
    ventas_rango = Venta.objects.filter(creado_en__gte=dt_inicio, creado_en__lte=dt_fin)
    total_ventas_rango = ventas_rango.aggregate(total=Sum('total'))['total'] or 0
    cantidad_ventas_rango = ventas_rango.count()
    
    # Ventas últimos 7 días dentro del rango
    ventas_7d = Venta.objects.filter(creado_en__gte=dt_hace_7, creado_en__lte=dt_fin)
    total_ventas_7d = ventas_7d.aggregate(total=Sum('total'))['total'] or 0
    
    # Ventas por estado
    ventas_por_estado = list(
        Venta.objects.filter(creado_en__gte=dt_inicio, creado_en__lte=dt_fin)
        .values('estado').annotate(cantidad=Count('id')).order_by('estado')
    )
    
    # Ventas por fecha (gráfico de línea)
    ventas_por_fecha = []
    dias = (fecha_fin - fecha_inicio).days + 1
    for i in range(dias):
        fecha = fecha_inicio + timedelta(days=i)
        # Crear rango datetime para cada día
        dt_dia_inicio = timezone.make_aware(datetime.combine(fecha, datetime.min.time()))
        dt_dia_fin = timezone.make_aware(datetime.combine(fecha, datetime.max.time()))
        total = Venta.objects.filter(creado_en__gte=dt_dia_inicio, creado_en__lte=dt_dia_fin).aggregate(total=Sum('total'))['total'] or 0
        ventas_por_fecha.append({
            'fecha': fecha.strftime('%a %d'),
            'total': float(total)
        })
    
    # Métodos de pago más usados
    metodos_pago = list(
        Venta.objects.filter(creado_en__gte=dt_inicio, creado_en__lte=dt_fin)
        .values('metodo_pago')
        .annotate(cantidad=Count('id'))
        .order_by('-cantidad')
    )
    
    # ===== CLIENTES =====
    total_clientes = Usuario.objects.filter(rol__nombre__iexact='Cliente').count()
    clientes_nuevos_rango = Usuario.objects.filter(
        rol__nombre__iexact='Cliente',
        creado_en__gte=dt_inicio, creado_en__lte=dt_fin
    ).count()
    
    # Clientes más activos (mayor cantidad de compras)
    clientes_top = list(
        Usuario.objects.filter(rol__nombre__iexact='Cliente')
        .annotate(num_compras=Count('venta'))
        .filter(num_compras__gt=0)
        .order_by('-num_compras')
        .values('nombre', 'num_compras', 'email')[:5]
    )
    
    # Productos más vendidos
    top_productos = list(
        VentaDetalle.objects.filter(venta__creado_en__gte=dt_inicio, venta__creado_en__lte=dt_fin)
        .values('producto__nombre')
        .annotate(total_vendido=Sum('cantidad'), ingresos=Sum('subtotal'))
        .order_by('-total_vendido')[:5]
    )
    
    # ===== HISTORIAL DE VENTAS DEL DÍA =====
    # Ventas de hoy
    dt_hoy_inicio = timezone.make_aware(datetime.combine(hoy, datetime.min.time()))
    dt_hoy_fin = timezone.make_aware(datetime.combine(hoy, datetime.max.time()))
    
    ventas_hoy = Venta.objects.filter(
        creado_en__gte=dt_hoy_inicio, 
        creado_en__lte=dt_hoy_fin
    ).select_related('usuario').prefetch_related('ventadetalle_set__producto').order_by('-creado_en')
    
    # Convertir a lista con detalles de detalles
    historial_ventas = []
    for venta in ventas_hoy:
        detalles = []
        for detalle in venta.ventadetalle_set.all():
            detalles.append({
                'producto': detalle.producto.nombre,
                'cantidad': detalle.cantidad,
                'precio': float(detalle.precio_unitario),
                'subtotal': float(detalle.subtotal)
            })
        historial_ventas.append({
            'id': venta.id,
            'cliente': venta.usuario.nombre if venta.usuario else 'Cliente Anónimo',
            'email': venta.usuario.email if venta.usuario else '',
            'hora': venta.creado_en.strftime('%H:%M:%S'),
            'metodo_pago': venta.metodo_pago,
            'estado': venta.estado,
            'total': float(venta.total),
            'detalles': detalles
        })
    
    return render(request, "panel/dashboard.html", {
        # Filtros
        "fecha_inicio": fecha_inicio.isoformat(),
        "fecha_fin": fecha_fin.isoformat(),
        
        # Inventario
        "total_prod": total_prod,
        "stock_total": stock_total,
        "productos_sin_stock": productos_sin_stock,
        "low_stock": low_stock[:8],
        "productos_por_categoria": productos_por_categoria,
        
        # Ventas
        "total_ventas_7d": float(total_ventas_7d),
        "total_ventas_rango": float(total_ventas_rango),
        "cantidad_ventas_rango": cantidad_ventas_rango,
        "ventas_por_fecha": ventas_por_fecha,
        "ventas_por_estado": ventas_por_estado,
        "metodos_pago": metodos_pago,
        
        # Clientes
        "total_clientes": total_clientes,
        "clientes_nuevos_rango": clientes_nuevos_rango,
        "clientes_top": clientes_top,
        
        # Productos más vendidos
        "top_productos": top_productos,
        
        # Rentabilidad
        "rentabilidad_categoria": rentabilidad_categoria,
        
        # Historial de ventas del día
        "historial_ventas": historial_ventas,
        "total_ventas_hoy": len(historial_ventas),
    })


# --------- Inventario / Productos ---------
@login_required
@group_required("Admin", "Empleado")
def inventario_list(request):
    q = request.GET.get("q", "").strip()
    cat = request.GET.get("categoria", "").strip()

    qs = (Producto.objects
          .select_related("categoria")
          .order_by("-id"))

    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
    if cat:
        qs = qs.filter(categoria_id=cat)

    categorias = Categoria.objects.all().order_by("nombre")
    return render(request, "panel/inventario_list.html", {
        "productos": qs,
        "categorias": categorias,
        "q": q,
        "categoria_sel": cat
    })


@login_required
@permission_required("productos.add_producto", raise_exception=True)
def producto_create(request):
    if request.method == "POST":
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            p = form.save()
            if p.imagen:
                messages.success(request, f"Producto «{p.nombre}» creado. Imagen guardada: {p.imagen.name}")
            else:
                messages.success(request, f"Producto «{p.nombre}» creado.")
            return redirect("panel:inventario_list")
        messages.error(request, "Revisa los errores del formulario.")
    else:
        form = ProductoForm()
    return render(request, "panel/producto_form.html", {"form": form, "modo": "Crear"})


@login_required
@permission_required("productos.change_producto", raise_exception=True)
def producto_update(request, pk):
    p = get_object_or_404(Producto, pk=pk)
    if request.method == "POST":
        form = ProductoForm(request.POST, request.FILES, instance=p)
        if form.is_valid():
            form.save()
            if p.imagen:
                messages.success(request, f"Producto «{p.nombre}» actualizado. Imagen guardada: {p.imagen.name}")
            else:
                messages.success(request, f"Producto «{p.nombre}» actualizado.")
            return redirect("panel:inventario_list")
        messages.error(request, "Revisa los errores del formulario.")
    else:
        form = ProductoForm(instance=p)
    return render(request, "panel/producto_form.html", {"form": form, "modo": "Editar"})


@login_required
@permission_required("productos.delete_producto", raise_exception=True)
def producto_delete(request, pk):
    p = get_object_or_404(Producto, pk=pk)
    if request.method == "POST":
        nombre = p.nombre
        p.delete()
        messages.success(request, f"Producto «{nombre}» eliminado.")
        return redirect("panel:inventario_list")
    return render(request, "panel/confirm_delete.html", {"obj": p, "tipo": "Producto"})


# --------- Categorías ---------
@login_required
@permission_required("productos.view_categoria", raise_exception=True)
def categoria_list(request):
    categorias = Categoria.objects.all().order_by("nombre")
    return render(request, "panel/categoria_list.html", {"categorias": categorias})


@login_required
@permission_required("productos.add_categoria", raise_exception=True)
def categoria_create(request):
    if request.method == "POST":
        form = CategoriaForm(request.POST)
        if form.is_valid():
            c = form.save()
            messages.success(request, f"Categoría «{c.nombre}» creada.")
            return redirect("panel:categoria_list")
        messages.error(request, "Corrige los errores del formulario.")
    else:
        form = CategoriaForm()
    return render(request, "panel/categoria_form.html", {"form": form, "modo": "Nueva"})


@login_required
@permission_required("productos.change_categoria", raise_exception=True)
def categoria_update(request, pk):
    c = get_object_or_404(Categoria, pk=pk)
    if request.method == "POST":
        form = CategoriaForm(request.POST, instance=c)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoría actualizada.")
            return redirect("panel:categoria_list")
        messages.error(request, "Corrige los errores del formulario.")
    else:
        form = CategoriaForm(instance=c)
    return render(request, "panel/categoria_form.html", {"form": form, "modo": "Editar"})


@login_required
@permission_required("productos.delete_categoria", raise_exception=True)
def categoria_delete(request, pk):
    c = get_object_or_404(Categoria, pk=pk)
    
    # Verificar si hay productos asociados
    productos_asociados = c.producto_set.count()
    
    if productos_asociados > 0:
        messages.error(request, f"No se puede eliminar la categoría «{c.nombre}» porque tiene {productos_asociados} producto(s) asociado(s). Primero debes asignar estos productos a otra categoría o eliminarlos.")
        return redirect("panel:categoria_list")
    
    if request.method == "POST":
        nombre = c.nombre
        c.delete()
        messages.success(request, f"Categoría «{nombre}» eliminada.")
        return redirect("panel:categoria_list")
    return render(request, "panel/confirm_delete.html", {"obj": c, "tipo": "Categoría"})


@login_required
def empleado_list(request):
    # Permitir siempre a superusers; para otros, comprobar permiso y mostrar mensaje en UI
    if not (request.user.is_superuser or request.user.has_perm('usuarios.view_usuario')):
        messages.error(request, 'No tienes permiso para ver la lista de empleados.')
        return redirect('panel:dashboard')
    qs = Usuario.objects.filter(rol__nombre__in=['Empleado', 'Administrador']).order_by('-creado_en')
    return render(request, 'panel/empleado_list.html', {'empleados': qs})


@login_required
def cliente_list(request):
    """Lista los clientes registrados (rol 'Cliente')."""
    # Permitir a superusers y a quien tenga permiso de ver usuarios
    if not (request.user.is_superuser or request.user.has_perm('usuarios.view_usuario')):
        messages.error(request, 'No tienes permiso para ver la lista de clientes.')
        return redirect('panel:dashboard')

    qs = Usuario.objects.filter(rol__nombre__iexact='Cliente').order_by('-creado_en')
    return render(request, 'panel/cliente_list.html', {'clientes': qs})


@login_required
@group_required("Admin", "Empleado")
def promociones_list(request):
    """Muestra productos cuya `fecha_vencimiento` esté dentro de los próximos 30 días
    y permite crear entradas en la tabla `promociones` seleccionando el tipo de promoción.
    """
    from django.utils import timezone
    from datetime import timedelta
    hoy = timezone.localdate()
    limite = hoy + timedelta(days=30)

    productos = Producto.objects.filter(fecha_vencimiento__isnull=False,
                                         fecha_vencimiento__range=(hoy, limite),
                                         estado='activo')

    if request.method == 'POST':
        # Procesar las promociones seleccionadas
        created = 0
        skipped = 0
        for p in productos:
            apply_key = f'apply_{p.pk}'
            if apply_key not in request.POST:
                continue

            key = f'promo_type_{p.pk}'
            tipo = request.POST.get(key)
            if not tipo or tipo == 'none':
                continue

            # Preparar campos para la tabla promociones
            nombre = f"Promo - {p.nombre}"
            descripcion = f"Promoción automática para producto que vence {p.fecha_vencimiento}. Tipo: {tipo}"
            descuento_val = None
            if tipo == 'descuento':
                # leer descuento opcional
                try:
                    descuento_val = request.POST.get(f'descuento_{p.pk}')
                    if descuento_val:
                        descuento_val = float(descuento_val)
                except Exception:
                    descuento_val = None

            # Rango de fechas: inicio hoy, fin el día de vencimiento
            fecha_inicio = hoy
            fecha_fin = p.fecha_vencimiento
            # Evitar duplicados:
            # Evitar duplicados: sólo comprobamos promociones internas vinculadas al mismo producto
            already_exists = False
            try:
                already_exists = Promotion.objects.filter(
                    producto=p,
                ).filter(
                    promotion_end__gte=fecha_inicio,
                    promotion_start__lte=fecha_fin,
                ).filter(status__in=['approved', 'pending']).exists()
            except Exception:
                already_exists = False

            if already_exists:
                skipped += 1
                continue

            # Crear registro en la tabla externa `promociones` (si existe) — no la usamos para bloqueo
            try:
                Promocion.objects.create(
                    nombre=nombre,
                    descripcion=descripcion,
                    descuento=descuento_val,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    activo='si'
                )
            except Exception:
                # No interrumpir por fallos en la tabla externa
                pass

            # Crear también la promoción interna ligada al producto
            try:
                # Mapear request.user a usuarios.Usuario si existe
                creador = None
                try:
                    creador = Usuario.objects.filter(email__iexact=request.user.email).first()
                except Exception:
                    creador = None

                # Server-side: si es descuento, forzar entero
                if tipo == 'descuento' and descuento_val is not None:
                    try:
                        descuento_val = int(float(descuento_val))
                    except Exception:
                        descuento_val = None

                Promotion.objects.create(
                    producto=p,
                    creado_por=creador,
                    tipo=tipo,
                    discount_percent=descuento_val,
                    recommended_reason=descripcion,
                    promotion_start=fecha_inicio,
                    promotion_end=fecha_fin,
                    status='approved' if tipo in ('2x1', 'descuento', 'oferta') else 'pending'
                )
            except Exception:
                # Ignorar errores individuales
                pass

            created += 1

        msgs = []
        if created:
            msgs.append(f'{created} promoción(es) creada(s).')
        if skipped:
            msgs.append(f'{skipped} promoción(es) omitida(s) por duplicados.')
        if msgs:
            messages.success(request, ' '.join(msgs))
        else:
            messages.info(request, 'No se crearon promociones. Selecciona al menos una.')
        return redirect('panel:promociones')

    # También listar promociones internas existentes para edición/eliminación
    try:
        # Obtener promociones y calcular días restantes hasta promotion_end (0 si ya expiró)
        promociones_existentes = list(Promotion.objects.select_related('producto').order_by('-creado_en'))
        for promo in promociones_existentes:
            dias = None
            try:
                if promo.promotion_end:
                    dias = (promo.promotion_end - hoy).days
            except Exception:
                dias = None
            # Normalizar: si dias es negativo considerarlo expirado (0 días restantes)
            if dias is None:
                promo.dias_restantes = None
            else:
                promo.dias_restantes = dias if dias >= 0 else 0
    except Exception:
        promociones_existentes = []

    return render(request, 'panel/promociones.html', {
        'productos': productos,
        'hoy': hoy,
        'limite': limite,
        'promociones_existentes': promociones_existentes,
    })


@login_required
@group_required("Admin", "Empleado")
def promociones_edit(request, pk):
    """Editar una promoción interna (modelo Promotion)."""
    p = get_object_or_404(Promotion, pk=pk)
    from .forms import PromotionForm
    if request.method == 'POST':
        form = PromotionForm(request.POST, instance=p)
        if form.is_valid():
            form.save()
            messages.success(request, 'Promoción actualizada.')
            return redirect('panel:promociones')
        messages.error(request, 'Corrige los errores del formulario.')
    else:
        form = PromotionForm(instance=p)
    return render(request, 'panel/promocion_edit.html', {'form': form, 'promocion': p})


@login_required
@group_required("Admin", "Empleado")
def promociones_delete(request, pk):
    """Eliminar una promoción interna."""
    p = get_object_or_404(Promotion, pk=pk)
    if request.method == 'POST':
        nombre = str(p)
        p.delete()
        messages.success(request, f'Promoción «{nombre}» eliminada.')
        return redirect('panel:promociones')
    return render(request, 'panel/confirm_delete.html', {'obj': p, 'tipo': 'Promoción'})


@login_required
@group_required("Admin", "Empleado")
def promociones_toggle(request, pk):
    """Alterna el estado de una promoción interna entre 'approved' y 'rejected'."""
    p = get_object_or_404(Promotion, pk=pk)
    # Solo aceptar POST para cambiar estado
    if request.method == 'POST':
        try:
            if p.status == 'approved':
                p.status = 'rejected'
                msg = 'Promoción deshabilitada.'
            else:
                p.status = 'approved'
                msg = 'Promoción habilitada.'
            p.save()
            messages.success(request, msg)
        except Exception as e:
            messages.error(request, f'Error al cambiar el estado: {e}')
        return redirect('panel:promociones')
    # Si se accede por GET mostrar confirmación simple
    return render(request, 'panel/confirm_toggle.html', {'obj': p})


@login_required
def empleado_create(request):
    # Comprobación explícita de permisos para dar feedback claro en UI
    if not request.user.has_perm('usuarios.add_usuario'):
        messages.error(request, 'No tienes permisos para crear empleados. Contacta al administrador.')
        return redirect('panel:dashboard')
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            try:
                usuario = form.save(commit=False)
                # Validar que no exista un Usuario con ese email
                if Usuario.objects.filter(email__iexact=usuario.email).exists():
                    form.add_error('email', 'Ya existe un usuario con este correo.')
                    messages.error(request, 'Ya existe un usuario con este correo.')
                    return render(request, 'panel/empleado_form.html', {'form': form, 'modo': 'Nuevo'})
                # Generar password aleatorio sencillo (se puede mejorar)
                import secrets
                raw_pw = secrets.token_urlsafe(8)
                # Generar un username único basado en el nombre
                base = ''.join(e for e in usuario.nombre.lower() if e.isalnum())
                if not base:
                    base = usuario.email.split('@')[0]
                username = base
                counter = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base}{counter}"
                    counter += 1

                # Crear auth.User con username generado y sin contraseña (se enviará link para crearla)
                auth_user = User.objects.create(username=username, email=usuario.email)
                auth_user.set_unusable_password()
                auth_user.save()

                # Asignar rol Empleado
                rol_emp, _ = Rol.objects.get_or_create(nombre='Empleado', defaults={'descripcion': 'Empleado del local'})
                usuario.rol = rol_emp
                # Forzar cambio de contraseña en el primer login
                usuario.must_change_password = True
                # Guardar contraseña hasheada en campo password
                from django.contrib.auth.hashers import make_password
                usuario.password = make_password(raw_pw)
                try:
                    usuario.save()
                except IntegrityError:
                    form.add_error('email', 'Ya existe un usuario con este correo.')
                    messages.error(request, 'Ya existe un usuario con este correo.')
                    return render(request, 'panel/empleado_form.html', {'form': form, 'modo': 'Nuevo'})

                # Asignar contraseña usable al auth_user y mostrar credenciales al admin (modo local)
                auth_user.set_password(raw_pw)
                auth_user.save()
                # Asegurar que el auth_user esté en el grupo Empleado
                try:
                    grupo_emp, _ = Group.objects.get_or_create(name='Empleado')
                    auth_user.groups.add(grupo_emp)
                except Exception:
                    pass
                messages.success(request, f'Empleado creado. Usuario: {username} | Contraseña: {raw_pw}')
                return redirect('panel:empleado_list')
            except Exception as e:
                # Mostrar el error en la UI para diagnóstico
                import traceback
                tb = traceback.format_exc()
                messages.error(request, f'Error al crear empleado: {e}')
                messages.error(request, f'Detalle: {tb.splitlines()[-1]}')
                return render(request, 'panel/empleado_form.html', {'form': form, 'modo': 'Nuevo'})
        messages.error(request, 'Corrige los errores del formulario.')
    else:
        form = UsuarioForm()
    return render(request, 'panel/empleado_form.html', {'form': form, 'modo': 'Nuevo'})


@login_required
def empleado_update(request, pk):
    if not (request.user.is_superuser or request.user.has_perm('usuarios.change_usuario')):
        messages.error(request, 'No tienes permiso para editar empleados.')
        return redirect('panel:dashboard')
    empleado = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=empleado)
        if form.is_valid():
            usuario = form.save(commit=False)
            # Si cambió el email, verificar no colisionar con otro Usuario
            if Usuario.objects.filter(email__iexact=usuario.email).exclude(pk=empleado.pk).exists():
                form.add_error('email', 'Otro usuario ya usa este correo.')
                messages.error(request, 'Otro usuario ya usa este correo.')
                return render(request, 'panel/empleado_form.html', {'form': form, 'modo': 'Editar'})
            # Sincronizar email en auth.User
            username = usuario.email.lower()
            auth_user, created = User.objects.get_or_create(username=username, defaults={'email': usuario.email})
            if auth_user.email != usuario.email:
                auth_user.email = usuario.email
                auth_user.save()
            # Si el admin está editando el empleado, asumimos que ya pasó el primer acceso
            usuario.must_change_password = False
            usuario.save()
            # Asegurar que el auth_user esté en el grupo Empleado
            try:
                grupo_emp, _ = Group.objects.get_or_create(name='Empleado')
                auth_user.groups.add(grupo_emp)
            except Exception:
                pass
            messages.success(request, 'Empleado actualizado.')
            return redirect('panel:empleado_list')
        messages.error(request, 'Corrige los errores del formulario.')
    else:
        form = UsuarioForm(instance=empleado)
    return render(request, 'panel/empleado_form.html', {'form': form, 'modo': 'Editar'})


@login_required
def empleado_delete(request, pk):
    if not (request.user.is_superuser or request.user.has_perm('usuarios.delete_usuario')):
        messages.error(request, 'No tienes permiso para eliminar empleados.')
        return redirect('panel:dashboard')
    empleado = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        nombre = empleado.nombre
        # Eliminar auth.User asociado si existe
        try:
            auth = User.objects.filter(username=empleado.email.lower()).first()
            if auth:
                auth.delete()
        except Exception:
            pass
        empleado.delete()
        messages.success(request, f'Empleado «{nombre}» eliminado.')
        return redirect('panel:empleado_list')
    return render(request, 'panel/confirm_delete.html', {'obj': empleado, 'tipo': 'Empleado'})


# --------- Cupones ---------
import string
import random

def generar_codigo_cupon():
    """Genera un código aleatorio de 6 caracteres con letras y números."""
    caracteres = string.ascii_uppercase + string.digits
    codigo = ''.join(random.choice(caracteres) for _ in range(6))
    return codigo


@login_required
@group_required("Admin", "Empleado")
def cupones_list(request):
    """Lista todos los cupones creados con opción de crear nuevos."""
    from .models import Cupon
    from decimal import Decimal
    
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')
        porcentaje = request.POST.get('porcentaje_descuento')
        
        # Validaciones
        if not producto_id or not porcentaje:
            messages.error(request, 'Debes seleccionar un producto y un porcentaje.')
            cupones = Cupon.objects.filter(is_deleted=False).select_related('producto', 'usuario')
            productos = Producto.objects.filter(estado='activo')
            return render(request, 'panel/cupones.html', {
                'cupones': cupones,
                'productos': productos,
            })
        
        try:
            porcentaje = int(porcentaje)
            if porcentaje < 1 or porcentaje > 50:
                messages.error(request, 'El porcentaje debe estar entre 1 y 50.')
                cupones = Cupon.objects.filter(is_deleted=False).select_related('producto', 'usuario')
                productos = Producto.objects.filter(estado='activo')
                return render(request, 'panel/cupones.html', {
                    'cupones': cupones,
                    'productos': productos,
                })
        except ValueError:
            messages.error(request, 'El porcentaje debe ser un número válido.')
            cupones = Cupon.objects.filter(is_deleted=False).select_related('producto', 'usuario')
            productos = Producto.objects.filter(estado='activo')
            return render(request, 'panel/cupones.html', {
                'cupones': cupones,
                'productos': productos,
            })
        
        try:
            producto = Producto.objects.get(pk=producto_id, estado='activo')
        except Producto.DoesNotExist:
            messages.error(request, 'El producto seleccionado no existe o no está activo.')
            cupones = Cupon.objects.filter(is_deleted=False).select_related('producto', 'usuario')
            productos = Producto.objects.filter(estado='activo')
            return render(request, 'panel/cupones.html', {
                'cupones': cupones,
                'productos': productos,
            })
        
        # Generar código único
        codigo = generar_codigo_cupon()
        while Cupon.objects.filter(codigo=codigo).exists():
            codigo = generar_codigo_cupon()
        
        # Calcular precio con descuento
        precio_original = producto.precio
        descuento_monto = (precio_original * Decimal(porcentaje)) / Decimal(100)
        precio_con_descuento = precio_original - descuento_monto
        
        # Crear cupón
        try:
            cupon = Cupon.objects.create(
                codigo=codigo,
                producto=producto,
                porcentaje_descuento=porcentaje,
                precio_original=precio_original,
                precio_con_descuento=precio_con_descuento,
                estado='Activo'
            )
            messages.success(request, f'Cupón «{codigo}» generado exitosamente para {producto.nombre}.')
        except Exception as e:
            messages.error(request, f'Error al crear el cupón: {str(e)}')
        
        return redirect('panel:cupones')
    
    # GET: Mostrar lista de cupones
    cupones = Cupon.objects.filter(is_deleted=False).select_related('producto', 'usuario')
    productos = Producto.objects.filter(estado='activo')
    
    return render(request, 'panel/cupones.html', {
        'cupones': cupones,
        'productos': productos,
    })


@login_required
@group_required("Admin", "Empleado")
def cupones_delete(request, pk):
    """Realizar eliminación lógica de un cupón."""
    from .models import Cupon
    cupon = get_object_or_404(Cupon, pk=pk)
    
    if request.method == 'POST':
        codigo = cupon.codigo
        cupon.is_deleted = True
        cupon.save()
        messages.success(request, f'Cupón «{codigo}» eliminado.')
        return redirect('panel:cupones')
    
    return render(request, 'panel/confirm_delete.html', {'obj': cupon, 'tipo': 'Cupón'})


# --------- EXPORTAR DASHBOARD ---------
@login_required
@group_required("Admin", "Empleado")
def export_dashboard_pdf(request):
    """Exportar dashboard como PDF."""
    from io import BytesIO
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from django.utils import timezone
    from datetime import timedelta
    from ventas.models import Venta, VentaDetalle
    
    # Obtener fechas del request
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    
    hoy = timezone.localdate()
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = timezone.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = timezone.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_inicio = hoy - timedelta(days=30)
            fecha_fin = hoy
    else:
        fecha_inicio = hoy - timedelta(days=30)
        fecha_fin = hoy
    
    # Obtener datos
    ventas_rango = Venta.objects.filter(creado_en__date__range=(fecha_inicio, fecha_fin))
    total_ventas = ventas_rango.aggregate(total=Sum('total'))['total'] or 0
    cantidad_ventas = ventas_rango.count()
    
    top_productos = list(
        VentaDetalle.objects.filter(venta__creado_en__date__range=(fecha_inicio, fecha_fin))
        .values('producto__nombre')
        .annotate(total_vendido=Sum('cantidad'), ingresos=Sum('subtotal'))
        .order_by('-total_vendido')[:5]
    )
    
    rentabilidad = list(
        VentaDetalle.objects.filter(venta__creado_en__date__range=(fecha_inicio, fecha_fin))
        .values('producto__categoria__nombre')
        .annotate(total_vendido=Sum('cantidad'), ingresos=Sum('subtotal'), num_ventas=Count('venta', distinct=True))
        .order_by('-ingresos')
    )
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#667eea'), spaceAfter=30)
    elements.append(Paragraph('REPORTE DE DASHBOARD', title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información de fechas
    info_text = f'<b>Período:</b> {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}'
    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumen general
    elements.append(Paragraph('<b>RESUMEN GENERAL</b>', styles['Heading2']))
    resumen_data = [
        ['Métrica', 'Valor'],
        ['Total de Ventas', f'${float(total_ventas):.2f}'],
        ['Cantidad de Transacciones', str(cantidad_ventas)],
        ['Promedio por Transacción', f'${float(total_ventas/cantidad_ventas) if cantidad_ventas > 0 else 0:.2f}'],
    ]
    resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Top productos
    elements.append(Paragraph('<b>TOP 5 PRODUCTOS MÁS VENDIDOS</b>', styles['Heading2']))
    top_data = [['Producto', 'Cantidad', 'Ingresos']]
    for prod in top_productos:
        top_data.append([prod['producto__nombre'], str(prod['total_vendido']), f"${prod['ingresos']:.2f}"])
    
    top_table = Table(top_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
    top_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(top_table)
    elements.append(PageBreak())
    
    # Rentabilidad por categoría
    elements.append(Paragraph('<b>ANÁLISIS DE RENTABILIDAD POR CATEGORÍA</b>', styles['Heading2']))
    rent_data = [['Categoría', 'Unidades', 'Ingresos', 'Ventas', 'Promedio']]
    for item in rentabilidad:
        cat = item['producto__categoria__nombre']
        unid = item['total_vendido']
        ingr = f"${item['ingresos']:.2f}"
        vent = item['num_ventas']
        prom = f"${item['ingresos']/item['num_ventas']:.2f}" if item['num_ventas'] > 0 else '$0.00'
        rent_data.append([cat, str(unid), ingr, str(vent), prom])
    
    rent_table = Table(rent_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1*inch, 1.2*inch])
    rent_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(rent_table)
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="dashboard_{fecha_inicio}_{fecha_fin}.pdf"'
    return response


@login_required
@group_required("Admin", "Empleado")
def export_dashboard_excel(request):
    """Exportar dashboard como Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    from datetime import timedelta
    from ventas.models import Venta, VentaDetalle
    
    # Obtener fechas del request
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    
    hoy = timezone.localdate()
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = timezone.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = timezone.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_inicio = hoy - timedelta(days=30)
            fecha_fin = hoy
    else:
        fecha_inicio = hoy - timedelta(days=30)
        fecha_fin = hoy
    
    # Obtener datos
    ventas_rango = Venta.objects.filter(creado_en__date__range=(fecha_inicio, fecha_fin))
    total_ventas = ventas_rango.aggregate(total=Sum('total'))['total'] or 0
    cantidad_ventas = ventas_rango.count()
    
    top_productos = list(
        VentaDetalle.objects.filter(venta__creado_en__date__range=(fecha_inicio, fecha_fin))
        .values('producto__nombre')
        .annotate(total_vendido=Sum('cantidad'), ingresos=Sum('subtotal'))
        .order_by('-total_vendido')[:5]
    )
    
    rentabilidad = list(
        VentaDetalle.objects.filter(venta__creado_en__date__range=(fecha_inicio, fecha_fin))
        .values('producto__categoria__nombre')
        .annotate(total_vendido=Sum('cantidad'), ingresos=Sum('subtotal'), num_ventas=Count('venta', distinct=True))
        .order_by('-ingresos')
    )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    
    # Estilos
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Título
    ws['A1'] = f"REPORTE DE DASHBOARD - {fecha_inicio} al {fecha_fin}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Resumen general
    row = 3
    ws[f'A{row}'] = "RESUMEN GENERAL"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    headers = ['Métrica', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row += 1
    data = [
        ['Total de Ventas', f"${float(total_ventas):.2f}"],
        ['Cantidad de Transacciones', cantidad_ventas],
        ['Promedio por Transacción', f"${float(total_ventas/cantidad_ventas) if cantidad_ventas > 0 else 0:.2f}"],
    ]
    for row_data in data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
        row += 1
    
    # Top productos
    row += 1
    ws[f'A{row}'] = "TOP 5 PRODUCTOS MÁS VENDIDOS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    headers = ['Producto', 'Cantidad', 'Ingresos']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row += 1
    for prod in top_productos:
        ws.cell(row=row, column=1).value = prod['producto__nombre']
        ws.cell(row=row, column=2).value = prod['total_vendido']
        ws.cell(row=row, column=3).value = f"${prod['ingresos']:.2f}"
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # Rentabilidad por categoría
    row += 2
    ws[f'A{row}'] = "ANÁLISIS DE RENTABILIDAD POR CATEGORÍA"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    headers = ['Categoría', 'Unidades', 'Ingresos', 'Num. Ventas', 'Promedio']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row += 1
    for item in rentabilidad:
        ws.cell(row=row, column=1).value = item['producto__categoria__nombre']
        ws.cell(row=row, column=2).value = item['total_vendido']
        ws.cell(row=row, column=3).value = f"${item['ingresos']:.2f}"
        ws.cell(row=row, column=4).value = item['num_ventas']
        ws.cell(row=row, column=5).value = f"${item['ingresos']/item['num_ventas']:.2f}" if item['num_ventas'] > 0 else '$0.00'
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Generar archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="dashboard_{fecha_inicio}_{fecha_fin}.xlsx"'
    wb.save(response)
    return response
